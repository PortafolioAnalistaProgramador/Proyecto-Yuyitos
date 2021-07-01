from django.shortcuts import render, redirect
from .models import CATEGORIA_PROVEEDOR, CLIENTE, FAMILIA_PRODUCTO, PROVEEDOR, PRODUCTO, ORDEN_PEDIDO, SEGUIMIENTO_PAGINA, TIPO_PRODUCTO, BOLETA, DETALLE_BOLETA, PAGO_FIADO, DETALLE_FIADO, DETALLE_ORDEN
from django.contrib import messages
from django import forms
from src.forms import (
    FormCliente, FormProveedor, FormProducto, 
    FormPedido, FormRegistroEdit, FormRegistroEdit2, 
    FormProveedorAct, FormFamiliaProd, FormProductoProv, 
    FormProductoEdit, FormClientesParaVenta, FormBoleta, 
    FormClientesInforme, FormInformeOrdenPedido, FormSeguimientoPagina, 
    FormCategProv, FormTipoProducto
)
import openpyxl
from tempfile import NamedTemporaryFile
from datetime import datetime
import numpy as np
from django.http import HttpResponse
from django.http import FileResponse
import webbrowser
import os
import wget
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, A2, A3
from reportlab.lib.pagesizes import mm
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT, TA_CENTER 
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, NextPageTemplate
from io import BytesIO
import re

from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.views.generic import ListView, DetailView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.messages.views import SuccessMessageMixin
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.contrib.auth.mixins import LoginRequiredMixin


def Login(request):

    username = request.POST.get('username')
    password = request.POST.get('password')

    if request.method == 'POST':
        user = authenticate(request, username = username, password = password)

        if user is not None:
            login(request, user)
            return render(request, 'index.html')
        else:
            messages.warning(request, 'Usuario o contraseña incorrectos')

    return render(request, 'registration/login.html')

@login_required(login_url="login")
def Index(request):
    return render(request, 'index.html')

@login_required(login_url="login")
def Venta(request):

    usuarioBoleta = request.user
    usuarioBoleta = User.objects.get(username=usuarioBoleta)

    seguimientoPag = SEGUIMIENTO_PAGINA.objects.create(
        pagina_visitada = "ventas",
        usuario = usuarioBoleta
    )
    seguimientoPag.save()

    productos = PRODUCTO.objects.all()
    form = FormClientesParaVenta()
    admin = User.objects.filter(username='Sra.Juanita')

    for juanita in admin:
        admin = juanita

    total = 0
    cliente = 0
    listaProductos = []
    cont = 0

    if request.method == 'POST':

        usuario = request.POST.get('usuario')
        contrasena = request.POST.get('contrasena')
        queryCliente = request.POST.get('cliente')
        fecha_pago = request.POST.get('fecha_pago')

        if queryCliente != '':

            if usuario == str(admin):

                user = authenticate(request, username = usuario, password = contrasena)
                
                if user is not None:
                    contador = 0
                    for key,value in request.POST.items():
                        contador = contador + 1

                    contador2 = 0
                    producto = []
                    for key,value in request.POST.items():

                        contador2 = contador2 + 1

                        if contador2 == contador:

                            total = value

                        if contador2 > 1 and contador2 < contador - 2:

                            if key == 'cliente':
                                cliente = value

                            if contador2 > 3:
                                cont += 1

                                producto.append(value)
                                
                                if cont == 3:
                                    listaProductos.append(producto)
                                    producto = []
                                    cont = 0

                    cliente = CLIENTE.objects.filter(id=cliente)
                    for client in cliente:
                        cliente = client

                    boleta = BOLETA.objects.create(
                        total_a_pagar = total,
                        usuario = usuarioBoleta,
                        cliente = cliente,
                        estado = 1
                    )
                    boleta.save()

                    bol = BOLETA.objects.all().last()
                    bol = BOLETA.objects.get(id = bol.id)
                    
                    for listaP in listaProductos:
                        
                        prod = PRODUCTO.objects.get(codigo_barra = listaP[0])

                        detalleBoleta = DETALLE_BOLETA.objects.create(
                            boleta = bol,
                            cantidad = listaP[1],
                            monto_a_pagar = listaP[2],
                            producto = prod
                        )
                        detalleBoleta.save()
                    
                    fecha_pago = fecha_pago[6:10] + '-' + fecha_pago[3:5] + '-' + fecha_pago[0:2]
                    pago_fiado = PAGO_FIADO.objects.create(
                        estado = 1,
                        monto = total,
                        fecha_final = fecha_pago,
                        cliente = CLIENTE.objects.get(id=queryCliente) 
                    )
                    

                    messages.warning(request, 'Venta realizada con exito')
                    return redirect('venta')

                else:
                    messages.warning(request, 'Usuario o contraseña incorrectos')
            else:
                messages.warning(request, 'Usuario ingresado no es administrador')

        else:
            if request.method == 'POST':
                usuarioBoleta = request.user
                usuarioBoleta = User.objects.get(username=usuarioBoleta)
                contador = 0

                for key,value in request.POST.items():
                    contador = contador + 1

                contador2 = 0
                producto = []
                for key,value in request.POST.items():

                    contador2 = contador2 +1

                    if contador2 == contador:

                        total = value

                    
                    if contador2 > 1 and contador2 < contador - 2:
                        
                        if contador2 > 3:
                            cont += 1

                            producto.append(value)

                            if cont == 3:
                                listaProductos.append(producto)
                                producto = []
                                cont = 0



                boleta = BOLETA.objects.create(
                    total_a_pagar = total,
                    usuario = usuarioBoleta,
                    estado = 1
                )
                boleta.save()

                bol = BOLETA.objects.all().last()
                bol = BOLETA.objects.get(id = bol.id)

                for listaP in listaProductos:
                    prod = PRODUCTO.objects.get(codigo_barra = listaP[0])

                    detalleBoleta = DETALLE_BOLETA.objects.create(
                        boleta = bol,
                        cantidad = listaP[1],
                        monto_a_pagar = listaP[2],
                        producto = prod
                    )
                    detalleBoleta.save()

                messages.warning(request, 'Venta realizada con exito')
                return redirect('venta')

    return render(request, 'venta.html',{'productos':productos, 'form':form})

@login_required(login_url="login")
def RecepcionPedido(request, id = None):

    ordenPedido = ORDEN_PEDIDO.objects.all()
    detalleOrden = DETALLE_ORDEN.objects.all()
    productos = ""

    ordenPedido2 = ""
    detalleOrden2 = ""
    if id:
        ordenPedido2 = ORDEN_PEDIDO.objects.filter(id=id)
        detalleOrden2 = DETALLE_ORDEN.objects.filter(orden_pedido=id)
        productos = PRODUCTO.objects.all()

    if request.method == 'POST':

        ordenPedido3 = ORDEN_PEDIDO.objects.get(id=id)
        validado = request.POST.get('validado')
        fecha = datetime.now().date()
        hora = datetime.now().time()
        print(fecha)
        print(hora)

        ordenPedido3.estado_recepcion = 1
        ordenPedido3.save()
        ordenPedido3.fecha_recepcion = fecha
        ordenPedido3.save()
        ordenPedido3.hora_recepcion = hora
        ordenPedido3.save()

        return redirect('RecepcionPedido')

    context= {
        'ordenPedido':ordenPedido,
        'productos':productos,
        'detalleOrden':detalleOrden,
        'ordenPedido2':ordenPedido2,
        'detalleOrden2':detalleOrden2
    }

    return render(request, 'recepcion_pedido.html',context)

# control de acceso para vistas basadas en clases
# @method_decorator(login_required, name='dispatch')
##**************************Usuarios***************************************
@login_required(login_url="login")
def UsuarioListado(request):

    usuarios = User.objects.all()

    return render(request, 'usuarios/listar.html', {'usuarios':usuarios})
    
@login_required(login_url="login")
def UsuarioCrear(request):

    username = request.POST.get('username')
    first_name = request.POST.get('first_name')
    last_name = request.POST.get('last_name')
    email = request.POST.get('email')
    password1 = request.POST.get('password1')
    password2 = request.POST.get('password2')
    is_superuser = request.POST.get('is_superuser')

    if is_superuser == 'on':
        is_superuser = True
    else:
        is_superuser = False

    validarEmail = User.objects.filter(email = email)
    validarUsuario = User.objects.filter(username = username)

    if request.method == 'POST':
        if validarEmail:
            messages.warning(request, 'el email ya existe')
        elif validarUsuario:
            messages.warning(request, 'el usuario ya existe')
        else:
            if (ValidacionCamposUsuario(request,
                username, first_name, last_name, email, password1, password2)
            ):
                username.strip()
                first_name.strip()
                last_name.strip()
                email.strip()
                password1.strip()
                password2.strip()

                user = User.objects.create_user(
                    username = username,
                    first_name = first_name,
                    last_name = last_name,
                    email = email,
                    is_superuser = is_superuser,
                    is_active = True
                    )
                user.set_password(password1)
                user.set_password(password2)

                if user is not None:
                    user.save()
                    messages.warning(request, 'Usuario creado correctamente')
                    return redirect('listarUsuarios')
                else:
                    messages.warning(request, 'No se pudo crear usuario')

    return render(request, 'usuarios/crear.html')

@login_required(login_url="login")
def UsuarioActualizar(request, id):

    user = User.objects.get(id=id)
    form = FormRegistroEdit(request.POST or None, instance=user)
    form2 = FormRegistroEdit2(request.POST or None, instance=user)
    username = request.POST.get('username')
    first_name = request.POST.get('first_name')
    last_name = request.POST.get('last_name')
    email = request.POST.get('email')
    password1 = request.POST.get('password1')
    password2 = request.POST.get('password2')
    is_superuser = request.POST.get('is_superuser')

    if is_superuser == 'on':
        is_superuser = True
    else:
        is_superuser = False

    if request.method == 'POST':
        if (ValidacionCamposUsuario(request,
            username, first_name, last_name, email, password1, password2)
        ):
            if user is not None:

                user.username = username.strip()
                user.save()
                user.first_name = first_name.strip()
                user.save()
                user.last_name = last_name.strip()
                user.save()
                user.email = email.strip()
                user.save()
                user.password1 = password1.strip()
                user.save()
                user.password2 = password2.strip()
                user.save()
                user.is_superuser = is_superuser
                user.save()
                messages.warning(request, 'Usuario actualizado correctamente')
                return redirect('listarUsuarios')
            else:
                messages.warning(request, 'No se pudo actualizar el usuario')
        else:
            messages.warning(request, 'No se pudo actualizar el usuario')

    return render(request, 'usuarios/actualizar.html', {'form':form, 'form2':form2})

def ValidacionCamposUsuario(request, username, first_name, last_name, email, password1, password2):
    estado = False
    username.strip()
    first_name.strip()
    last_name.strip()
    email.strip()
    password1.strip()
    password2.strip()

    if len(username) < 4:
        messages.warning(request, 'La cantidad de caracteres del usuario debe ser mayor a 3')
    elif len(username) > 150:
        messages.warning(request, 'La cantidad de caracteres del usuario debe ser menor a 151')
    elif len(first_name) < 3:
        messages.warning(request, 'La cantidad de caracteres del nombre debe ser mayor a 2')
    elif len(first_name) > 150:
        messages.warning(request, 'La cantidad de caracteres del nombre debe ser menor a 151')
    elif len(last_name) < 3:
        messages.warning(request, 'La cantidad de caracteres del apellido debe ser mayor a 2')
    elif len(last_name) > 150:
        messages.warning(request, 'La cantidad de caracteres del nombre debe ser menor a 151')
    elif len(email) < 4:
        messages.warning(request, 'La cantidad de caracteres del email debe ser mayor a 3')
    elif len(email) > 254:
        messages.warning(request, 'La cantidad de caracteres del email debe ser menor a 255')
    elif len(password1) < 4:
        messages.warning(request, 'La cantidad de caracteres de la contraseña debe ser mayor a 3')
    elif len(password1) > 128:
        messages.warning(request, 'La cantidad de caracteres de la contraseña debe ser menor a 129')
    elif password2 != password1:
        messages.warning(request, 'Las contraseñas tienen que ser iguales')
    else:
        if re.match('^[(a-z0-9\_\-\.)]+@[(a-z0-9\_\-\.)]+\.[(a-z)]{2,15}$',email.lower()):
            estado = True
        else:
            messages.warning(request, 'El correo no cumple con la estructura basica')

    return estado

@login_required(login_url="login")
def UsuarioDetalle(request, id):

    usuario = User.objects.get(id=id)

    return render(request, 'usuarios/detalles.html', {'usuario':usuario})

def DesactivarUsuario(request, id):
    user = User.objects.get(id = id)
    user.is_active = False
    user.save()
    messages.warning(request, f'Usuario {user.first_name} desactivado')
    return redirect('listarUsuarios')

def ActivarUsuario(request, id):
    user = User.objects.get(id = id)
    user.is_active = True
    user.save()
    messages.warning(request, f'Usuario {user.first_name} activado')
    return redirect('listarUsuarios')

##*************************************************************************

##**************************Clientes***************************************
@login_required(login_url="login")
def ClienteListado(request):

    clientes = CLIENTE.objects.all()

    return render(request, 'clientes/listar.html', {'clientes':clientes})

@login_required(login_url="login")
def ClienteCrear(request):

    run = request.POST.get('run')
    nombre = request.POST.get('nombre')
    telefono = request.POST.get('telefono')
    correo = request.POST.get('correo')
    direccion = request.POST.get('direccion')
    validarRun = CLIENTE.objects.filter(run=run)

    if request.method == 'POST':
        if validarRun:
            messages.warning(request, 'el run ya existe')
        else:
            if (ValidacionCamposCliente(request,
                run, nombre, telefono, correo, direccion)
            ):
                cliente = CLIENTE.objects.create(
                    run = run.strip(),
                    nombre = nombre.strip(),
                    telefono = telefono.strip(),
                    correo = correo.strip(),
                    direccion = direccion.strip(),
                    estado = 1
                )

                if cliente is not None:
                    cliente.save()
                    messages.warning(request, 'Cliente creado correctamente')
                    return redirect('listarClientes')
                else:
                    messages.warning(request, 'No se pudo crear el cliente')
            else:
                messages.warning(request, 'No se pudo crear el cliente')

    return render(request, 'clientes/crear.html')

@login_required(login_url="login")
def ClienteActualizar(request, id):

    cliente = CLIENTE.objects.get(id=id)
    form = FormCliente(request.POST or None, instance=cliente)

    run = request.POST.get('run')
    nombre = request.POST.get('nombre')
    telefono = request.POST.get('telefono')
    correo = request.POST.get('correo')
    direccion = request.POST.get('direccion')
    validarRun = CLIENTE.objects.filter(run=run)

    if request.method == 'POST':
        if (ValidacionCamposCliente(request,
            run, nombre, telefono, correo, direccion)
        ):
            if cliente is not None:

                cliente.run = run.strip()
                cliente.save()
                cliente.nombre = nombre.strip()
                cliente.save()
                cliente.telefono = telefono.strip()
                cliente.save()
                cliente.correo = correo.strip()
                cliente.save()
                cliente.direccion = direccion.strip()
                cliente.save()
                messages.warning(request, 'Cliente actualizado correctamente')
                return redirect('listarClientes')
            else:
                messages.warning(request, 'No se pudo actualizar el cliente')
        else:
            messages.warning(request, 'No se pudo actualizar el cliente')

    return render(request, 'clientes/actualizar.html', {'form':form})

def ValidacionCamposCliente(request, run, nombre, telefono, correo, direccion):

    estado = False
    run.strip()
    nombre.strip()
    telefono.strip()
    correo.strip()
    direccion.strip()

    if len(run) < 7:
        messages.warning(request, 'La cantidad de caracteres del run debe ser mayor a 6')
    elif len(run) > 10:
        messages.warning(request, 'La cantidad de caracteres del run debe ser menor a 11')
    elif len(nombre) < 3:
        messages.warning(request, 'La cantidad de caracteres del nombre debe ser mayor a 2')
    elif len(nombre) > 100:
        messages.warning(request, 'La cantidad de caracteres del nombre debe ser menor a 101')
    elif len(telefono) < 5:
        messages.warning(request, 'La cantidad de caracteres del telefono debe ser mayor a 4')
    elif len(telefono) > 12:
        messages.warning(request, 'La cantidad de caracteres del telefono debe ser menor a 13')
    elif len(correo) < 5:
        messages.warning(request, 'La cantidad de caracteres del correo debe ser mayor a 4')
    elif len(correo) > 128:
        messages.warning(request, 'La cantidad de caracteres del correo debe ser menor a 129')
    elif len(direccion) < 3:
        messages.warning(request, 'La cantidad de caracteres de la direccion debe ser mayor a 2')
    elif len(direccion) > 150:
        messages.warning(request, 'La cantidad de caracteres de la direccion debe ser menor a 151')
    else:
        if re.match('^[(a-z0-9\_\-\.)]+@[(a-z0-9\_\-\.)]+\.[(a-z)]{2,15}$',correo.lower()):
            estado = True
        else:
            messages.warning(request, 'El correo no cumple con la estructura basica')

    return estado

@login_required(login_url="login")
def ClienteDetalle(request, id):

    cliente = CLIENTE.objects.get(id=id)
    return render(request, 'clientes/detalles.html', {'cliente':cliente})

def DesactivarCliente(request, id):
    cliente = CLIENTE.objects.get(id = id)
    cliente.estado = 0
    cliente.save()
    messages.warning(request, f'Cliente {cliente.nombre} desactivado')
    return redirect('listarClientes')

def ActivarCliente(request, id):
    cliente = CLIENTE.objects.get(id = id)
    cliente.estado = 1
    cliente.save()
    messages.warning(request, f'Cliente {cliente.nombre} activado')
    return redirect('listarClientes')

##******************************************************************

##**************************Proveedor***************************************
@login_required(login_url="login")
def ProveedorListado(request):

    proveedores = PROVEEDOR.objects.all()
    return render(request, 'proveedores/listar.html', {'proveedores':proveedores})

@login_required(login_url="login")
def ProveedorCrear(request):

    form = FormProveedor(request.POST)
    razon_social = request.POST.get('razon_social')
    correo = request.POST.get('correo')
    telefono = request.POST.get('telefono')
    direccion = request.POST.get('direccion')
    categoria = request.POST.get('categoria_proveedor')
    categoria = CATEGORIA_PROVEEDOR.objects.filter(id = categoria)
    for cat in categoria:
        categoria = cat

    if request.method == 'POST':

        if (ValidacionCamposProveedor(request,
            razon_social, correo, telefono, direccion)
        ):
            proveedor = PROVEEDOR.objects.create(
                razon_social = razon_social.strip(),
                correo = correo.strip(),
                telefono = telefono.strip(),
                direccion = direccion.strip(),
                categoria_proveedor = categoria,
                estado = 1
            )

            if proveedor is not None:
                proveedor.save()
                messages.warning(request, 'Proveedor creado correctamente')
                return redirect('listarProveedores')
            else:
                messages.warning(request, 'No se pudo crear el Proveedor')

    return render(request, 'proveedores/crear.html',{'form':form})

@login_required(login_url="login")
def ProveedorActualizar(request, id):

    proveedor = PROVEEDOR.objects.get(id=id)
    form = FormProveedorAct(request.POST or None, instance=proveedor)
    razon_social = request.POST.get('razon_social')
    correo = request.POST.get('correo')
    telefono = request.POST.get('telefono')
    direccion = request.POST.get('direccion')

    categoria = request.POST.get('categoria_proveedor')
    categoria = CATEGORIA_PROVEEDOR.objects.filter(id = categoria)
    for cat in categoria:
        categoria = cat

    if request.method == 'POST':
        if (ValidacionCamposProveedor(request,
            razon_social, correo, telefono, direccion)
        ):
            if proveedor is not None:

                proveedor.razon_social = razon_social.strip()
                proveedor.save()
                proveedor.correo = correo.strip()
                proveedor.save()
                proveedor.telefono = telefono.strip()
                proveedor.save()
                proveedor.direccion = direccion.strip()
                proveedor.save()
                proveedor.categoria_proveedor = categoria
                proveedor.save()

                messages.warning(request, 'Proveedor actualizado correctamente')
                return redirect('listarProveedores')
            else:
                messages.warning(request, 'No se pudo actualizar el proveedor')
        else:
            messages.warning(request, 'No se pudo actualizar el proveedor')

    return render(request, 'proveedores/actualizar.html', {'form':form})

def ValidacionCamposProveedor(request, razon_social, correo, telefono, direccion):
    estado = False
    razon_social.strip()
    correo.strip()
    telefono.strip()
    direccion.strip()

    if len(razon_social) < 3:
        messages.warning(request, 'La cantidad de caracteres de la razon social debe ser mayor a 2')
    elif len(razon_social) > 100:
        messages.warning(request, 'La cantidad de caracteres de la razon social debe ser menor a 11')
    elif len(correo) < 3:
        messages.warning(request, 'La cantidad de caracteres del correo debe ser mayor a 2')
    elif len(correo) > 100:
        messages.warning(request, 'La cantidad de caracteres del correo debe ser menor a 101')
    elif len(telefono) < 5:
        messages.warning(request, 'La cantidad de caracteres del telefono debe ser mayor a 4')
    elif len(telefono) > 12:
        messages.warning(request, 'La cantidad de caracteres del telefono debe ser menor a 13')
    elif len(direccion) < 3:
        messages.warning(request, 'La cantidad de caracteres de la direccion debe ser mayor a 2')
    elif len(direccion) > 150:
        messages.warning(request, 'La cantidad de caracteres de la direccion debe ser menor a 151')
    else:
        if re.match('^[(a-z0-9\_\-\.)]+@[(a-z0-9\_\-\.)]+\.[(a-z)]{2,15}$',correo.lower()):
            estado = True
        else:
            messages.warning(request, 'El correo no cumple con la estructura basica')

    return estado

@login_required(login_url="login")
def ProveedorDetalle(request, id):
    
    proveedor = PROVEEDOR.objects.get(id=id)
    return render(request, 'proveedores/detalles.html', {'proveedor':proveedor})

def DesactivarProveedor(request, id):
    proveedor = PROVEEDOR.objects.get(id = id)
    proveedor.estado = 0
    proveedor.save()
    messages.warning(request, f'Proveedor {proveedor.razon_social} desactivado')
    return redirect('listarProveedores')

def ActivarProveedor(request, id):
    proveedor = PROVEEDOR.objects.get(id = id)
    proveedor.estado = 1
    proveedor.save()
    messages.warning(request, f'Cliente {proveedor.razon_social} activado')
    return redirect('listarProveedores')

##******************************************************************

##**************************Producto***************************************
@login_required(login_url="login")
def ProductoListado(request):
    
    productos = PRODUCTO.objects.all()
    return render(request, 'productos/listar.html', {'productos':productos})

@login_required(login_url="login")
def ProductoCrear(request):

    formProd = FormProducto(request.POST)
    formFam = FormFamiliaProd(request.POST)
    formProv = FormProductoProv(request.POST)
    nombre = request.POST.get('nombre')
    precio = request.POST.get('precio')
    precio_compra = request.POST.get('precio_compra')
    descripcion = request.POST.get('descripcion')
    stock = request.POST.get('stock')
    stock_critico = request.POST.get('stock_critico')
    proveedor = request.POST.get('proveedor')
    familia_producto = request.POST.get('familia_producto')
    fecha_vencimiento = request.POST.get('fecha_vencimiento')
    hora_vencimiento = request.POST.get('hora_vencimiento')
    tipo_producto = request.POST.get('tipo_producto')

    if request.method == 'POST':

        if (ValidacionCamposProducto(request,
            nombre, precio, descripcion, precio_compra, stock, stock_critico, proveedor, familia_producto, tipo_producto)
        ):
            if fecha_vencimiento:
                fecha_v = fecha_vencimiento[6:10] + '-' + fecha_vencimiento[3:5] + '-' + fecha_vencimiento[0:2] + ' ' + hora_vencimiento + ':00'
                fecha_vencimiento = fecha_vencimiento[0:2]+fecha_vencimiento[3:5]+fecha_vencimiento[6:10]
            else:
                fecha_v = "1000-10-10 00:00:00"
                fecha_vencimiento = "00000000"

            familia_producto = str(familia_producto).zfill(3)
            tipo_producto = str(tipo_producto).zfill(3)
            proveedor = str(proveedor).zfill(3)

            codigo_barra = proveedor + familia_producto + fecha_vencimiento + tipo_producto

            if familia_producto:
                familia_producto = FAMILIA_PRODUCTO.objects.filter(id = familia_producto)
            for famP in familia_producto:
                familia_producto = famP

            producto = PRODUCTO.objects.create(
                nombre = nombre.strip(),
                precio = precio.strip(),
                descripcion = descripcion.strip(),
                precio_compra = precio_compra.strip(),
                stock = stock.strip(),
                stock_critico = stock_critico.strip(),
                estado = 1,
                fecha_vencimiento = fecha_v,
                codigo_barra = codigo_barra,
                familia_producto = familia_producto
            )

            if producto is not None:
                producto.save()
                messages.warning(request, 'Producto creado correctamente')
                return redirect('listarProductos')
            else:
                messages.warning(request, 'No se pudo crear el Producto')

    return render(request, 'productos/crear.html',{'formProd':formProd, 'formFam':formFam, 'formProv':formProv})

@login_required(login_url="login")
def ProductoActualizar(request, id):

    producto = PRODUCTO.objects.get(id=id)

    form = FormProductoEdit(request.POST or None, instance=producto)
    formProd = FormProducto(request.POST or None, instance=producto)

    tipo_producto = len(producto.codigo_barra)
    tipo_producto = int(producto.codigo_barra[-3:tipo_producto])
    tipo_producto = TIPO_PRODUCTO.objects.get(id=tipo_producto)

    proveedor = int(producto.codigo_barra[0:3])
    proveedor = PROVEEDOR.objects.get(id=proveedor)

    fecha = str(producto.fecha_vencimiento)
    hora = str(producto.fecha_vencimiento)
    fecha = fecha[0:10]
    hora = hora[11:16]

    formFam = FormFamiliaProd(request.POST or None, instance=tipo_producto)
    formProv = FormProductoProv(request.POST or None, instance=proveedor)

    context = {
        'formProd':formProd,
        'formFam':formFam,
        'formProv':formProv,
        'form':form,
        'tipo_producto':tipo_producto,
        'proveedor':proveedor,
        'fecha':fecha,
        'hora':hora
    }

    nombre = request.POST.get('nombre')
    precio = request.POST.get('precio')
    precio_compra = request.POST.get('precio_compra')
    descripcion = request.POST.get('descripcion')
    stock = request.POST.get('stock')
    stock_critico = request.POST.get('stock_critico')
    proveedor = request.POST.get('proveedor')
    familia_producto = request.POST.get('familia_producto')
    fecha_vencimiento = request.POST.get('fecha_vencimiento')
    hora_vencimiento = request.POST.get('hora_vencimiento')
    tipo_producto = request.POST.get('tipo_producto')

    if request.method == 'POST':
        if (ValidacionCamposProducto(request,
                nombre, precio, descripcion, precio_compra, stock, stock_critico, proveedor, familia_producto, tipo_producto)
            ):
                if fecha_vencimiento:
                    fecha_v = fecha_vencimiento[6:10] + '-' + fecha_vencimiento[3:5] + '-' + fecha_vencimiento[0:2] + ' ' + hora_vencimiento + ':00'
                    fecha_vencimiento = fecha_vencimiento[0:2]+fecha_vencimiento[3:5]+fecha_vencimiento[6:10]
                else:
                    fecha_v = "1000-10-10 00:00:00"
                    fecha_vencimiento = "00000000"

                familia_producto = str(familia_producto).zfill(3)
                tipo_producto = str(tipo_producto).zfill(3)
                proveedor = str(proveedor).zfill(3)

                codigo_barra = proveedor + familia_producto + fecha_vencimiento + tipo_producto

                if familia_producto:
                    familia_producto = FAMILIA_PRODUCTO.objects.filter(id = familia_producto)
                for famP in familia_producto:
                    familia_producto = famP

                if producto is not None:
                    producto.nombre = nombre.strip()
                    producto.save()

                    producto.precio = precio.strip()
                    producto.save()

                    producto.descripcion = descripcion.strip()
                    producto.save()

                    producto.precio_compra = precio_compra.strip()
                    producto.save()
                    producto.stock = stock.strip()
                    producto.save()
                    producto.stock_critico = stock_critico.strip()
                    producto.save()
                    producto.estado = 1
                    producto.save()
                    producto.fecha_vencimiento = fecha_v
                    producto.save()
                    producto.codigo_barra = codigo_barra
                    producto.save()
                    producto.familia_producto = familia_producto
                    producto.save()

                    messages.warning(request, 'Producto creado correctamente')
                    return redirect('listarProductos')
                else:
                    messages.warning(request, 'No se pudo crear el Producto')

    return render(request, 'productos/actualizar.html',context)

def ValidacionCamposProducto(
    request,
    nombre,
    precio,
    descripcion,
    precio_compra,
    stock,
    stock_critico,
    proveedor,
    familia_producto,
    tipo_producto
):
    estado = False
    nombre.strip()
    precio.strip()
    descripcion.strip()
    precio_compra.strip()
    stock.strip()
    stock_critico.strip()

    if len(nombre) < 3:
        messages.warning(request, 'La cantidad de caracteres del nombre debe ser mayor a 2')
    elif len(nombre) > 100:
        messages.warning(request, 'La cantidad de caracteres del nombre debe ser menor a 101')
    elif len(precio) < 1:
        messages.warning(request, 'La cantidad de caracteres del precio debe ser mayor a 0')
    elif len(precio) > 10:
        messages.warning(request, 'La cantidad de caracteres del precio debe ser menor a 11')
    elif len(descripcion) < 4:
        messages.warning(request, 'La cantidad de caracteres de la descripcion debe ser mayor a 3')
    elif len(descripcion) > 200:
        messages.warning(request, 'La cantidad de caracteres de la descripcion debe ser menor a 201')
    elif len(precio_compra) < 1:
        messages.warning(request, 'La cantidad de caracteres del precio de compra debe ser mayor a 0')
    elif len(precio_compra) > 10:
        messages.warning(request, 'La cantidad de caracteres del precio de compra debe ser menor a 11')
    elif len(stock) < 1:
        messages.warning(request, 'La cantidad de caracteres del stock debe ser mayor a 0')
    elif len(stock) > 10:
        messages.warning(request, 'La cantidad de caracteres del stock debe ser menor a 11')
    elif len(stock_critico) < 1:
        messages.warning(request, 'La cantidad de caracteres del stock critico debe ser mayor a 0')
    elif len(stock_critico) > 10:
        messages.warning(request, 'La cantidad de caracteres del stock critico debe ser menor a 11')
    elif len(proveedor) == 0:
        messages.warning(request, 'Se debe seleccionar al menos una opcion en proveedor')
    elif len(familia_producto) == 0:
        messages.warning(request, 'Se debe seleccionar al menos una opcion en familia de producto')
    elif len(tipo_producto) == 0:
         messages.warning(request, 'Se debe seleccionar al menos una opcion en tipo de producto')
    else:
        estado = True

    return estado

@login_required(login_url="login")
def ProductoDetalle(request, id):

    producto = PRODUCTO.objects.get(id=id)
    return render(request, 'productos/detalles.html', {'producto':producto})

def DesactivarProducto(request, id):
    producto = PRODUCTO.objects.get(id = id)
    producto.estado = 0
    producto.save()
    messages.warning(request, f'Producto {producto.nombre} desactivado')
    return redirect('listarProductos')

def ActivarProducto(request, id):
    producto = PRODUCTO.objects.get(id = id)
    producto.estado = 1
    producto.save()
    messages.warning(request, f'Producto {producto.nombre} activado')
    return redirect('listarProductos')

##********************************************************************

##**************************Tipo de producto********************
@login_required(login_url="login")
def TipoProductoListado(request):

    tipos = TIPO_PRODUCTO.objects.all()
    return render(request, 'tipos_productos/listar.html', {'tipos':tipos})

@login_required(login_url="login")
def TipoProductoCrear(request):
    
    form = FormTipoProducto()

    if request.method == 'POST':

        descripcion = request.POST.get('descripcion')
        proveedor = request.POST.get('proveedor')
        
        if len(descripcion) <= 3:
            messages.warning(request, 'La descripcion no puede tener una cantidad de caracteres menor a 3')
        elif proveedor == "":
            messages.warning(request, 'En proveedor debe seleccionar alguno de la lista')
        else:
            proveedor = PROVEEDOR.objects.get(id = proveedor)

            tipo_producto = TIPO_PRODUCTO.objects.create(
                descripcion = descripcion,
                proveedor = proveedor,
            )

            if tipo_producto is not None:
                tipo_producto.save()
                messages.warning(request, 'Tipo producto creado correctamente')
                return redirect('listarTiposProductos')
            else:
                messages.warning(request, 'No se pudo crear el tipo producto')

    return render(request, 'tipos_productos/crear.html',{'form':form})

@login_required(login_url="login")
def TipoProductoActualizar(request, id):
    
    tipo_producto = TIPO_PRODUCTO.objects.get(id=id)
    form = FormTipoProducto(request.POST or None, instance=tipo_producto)
    
    
    if request.method == 'POST':

        descripcion = request.POST.get('descripcion')
        proveedor = request.POST.get('proveedor')

        if len(descripcion) <= 3:
            messages.warning(request, 'La descripcion no puede tener una cantidad de caracteres menor a 3')
        elif proveedor == "":
            messages.warning(request, 'En proveedor debe seleccionar alguno de la lista')
        elif tipo_producto is not None:
            proveedor = PROVEEDOR.objects.get(id = proveedor)
            tipo_producto.descripcion = descripcion
            tipo_producto.save()
            tipo_producto.proveedor = proveedor
            tipo_producto.save()

            messages.warning(request, 'Tipo producto actualizado correctamente')
            return redirect('listarTiposProductos')
        else:
            messages.warning(request, 'No se pudo actualizar el tipo producto')
        
    return render(request, 'tipos_productos/actualizar.html',{'form':form})

##******************************************************************

##**************************Familia de producto********************
@login_required(login_url="login")
def FamiliaProductoListado(request):

    familiaP = FAMILIA_PRODUCTO.objects.all()
    return render(request, 'familia_producto/listar.html', {'familiaP':familiaP})

@login_required(login_url="login")
def FamiliaProductoCrear(request):
    
    form = FormFamiliaProd()

    if request.method == 'POST':

        descripcion = request.POST.get('descripcion')
        tipo_producto = request.POST.get('tipo_producto')
        
        if len(descripcion) <= 3:
            messages.warning(request, 'La descripcion no puede tener una cantidad de caracteres menor a 3')
        elif tipo_producto == "":
            messages.warning(request, 'En tipo producto debe seleccionar alguno de la lista')
        else:
            tipo_producto = TIPO_PRODUCTO.objects.get(id = tipo_producto)
            
            familia_producto = FAMILIA_PRODUCTO.objects.create(
                descripcion = descripcion,
                tipo_producto = tipo_producto,
            )

            if familia_producto is not None:
                familia_producto.save()
                messages.warning(request, 'Familia producto creado correctamente')
                return redirect('listarFamiliasProductos')
            else:
                messages.warning(request, 'No se pudo crear el familia producto')

    return render(request, 'familia_producto/crear.html',{'form':form})
   
@login_required(login_url="login")
def FamiliaProductoActualizar(request, id):
    
    familia_producto = FAMILIA_PRODUCTO.objects.get(id=id)
    form = FormFamiliaProd(request.POST or None, instance=familia_producto)
    
    if request.method == 'POST':

        descripcion = request.POST.get('descripcion')
        tipo_producto = request.POST.get('tipo_producto')

        if len(descripcion) <= 3:
            messages.warning(request, 'La descripcion no puede tener una cantidad de caracteres menor a 3')
        elif tipo_producto == "":
            messages.warning(request, 'En proveedor debe seleccionar alguno de la lista')
        elif familia_producto is not None:
            tipo_producto = TIPO_PRODUCTO.objects.get(id = tipo_producto)

            familia_producto.descripcion = descripcion
            familia_producto.save()
            familia_producto.tipo_producto = tipo_producto
            familia_producto.save()

            messages.warning(request, 'Familia producto actualizado correctamente')
            return redirect('listarFamiliasProductos')
        else:
            messages.warning(request, 'No se pudo actualizar la familia producto')
        
    return render(request, 'familia_producto/actualizar.html',{'form':form})

##******************************************************************

##**************************Pedidos********************
@login_required(login_url="login")
def PedidosListado(request):
    
    ordenP = ORDEN_PEDIDO.objects.all()
    return render(request, 'pedidos/listar.html', {'ordenP':ordenP})

def PedidosCrear(request, id = None):

    proveedores = PROVEEDOR.objects.all()

    tiposProductos = TIPO_PRODUCTO.objects.filter(proveedor=id)

    listaF = []
    for tiposP in tiposProductos:
        famId = FAMILIA_PRODUCTO.objects.filter(tipo_producto=tiposP)
        for f in famId:
            listaF.append(f.id)

    productos = PRODUCTO.objects.all()

    listaProds = []
    for prod in productos:
        idFam = FAMILIA_PRODUCTO.objects.filter(descripcion=prod.familia_producto)

        for idF in idFam:

            for listF in listaF:
                if listF == idF.id:

                    listaProds.append(prod.nombre)

    if request.method == 'POST':

        listaProductos = []
        producto = []

        contador2 = 0
        for key,value in request.POST.items():
            contador2 += 1

        contador = 0
        fecha = ""
        cont = 0
        for key,value in request.POST.items():
            
            contador += 1
            if contador == 2:
                proveedorOrden = int(value)

            if contador > 2:

                if contador == contador2:
                    if value == "":
                        fecha = "1000-10-10"
                    else:
                        fecha = value
                        fecha = fecha[6:10]+ '-' +fecha[3:5]+ '-' + fecha[0:2]
                        print(fecha)

                if contador < contador2:
                    cont += 1

                    producto.append(value)

                    if cont == 2:
                        listaProductos.append(producto)
                        producto = []
                        cont = 0

        proveedorOrden = PROVEEDOR.objects.get(id=proveedorOrden)

        ordenPedido = ORDEN_PEDIDO.objects.create(
            estado_recepcion = 0,
            proveedor = proveedorOrden,
            fecha_llegada = fecha
        )
        ordenPedido.save()

        ordenPedido = ORDEN_PEDIDO.objects.all().last()
        ordenPedido = ORDEN_PEDIDO.objects.get(id = ordenPedido.id)

        for listaP in listaProductos:
            
            prod = PRODUCTO.objects.get(nombre=listaP[0])

            detallePedido = DETALLE_ORDEN.objects.create(
                producto = prod,
                cantidad = listaP[1],
                orden_pedido = ordenPedido
            )
            detallePedido.save()

        messages.warning(request, 'Orden de pedido realizada con exito')
        return redirect('listarPedidos')


    context = {
        'proveedores':proveedores,
        'listaProds':listaProds
    }
    return render(request, 'pedidos/crear.html', context)

@login_required(login_url="login")
def PedidosDetalle(request, id):
    pedido = ORDEN_PEDIDO.objects.filter(id=id)
    detallesP = DETALLE_ORDEN.objects.filter(orden_pedido=id)
    context = {
        'pedido':pedido,
        'detallesP':detallesP,
        'id':id
    }
    return render(request, 'pedidos/detalles.html', context)

def DesactivarPedido(request, id):
    pedido = ORDEN_PEDIDO.objects.get(id = id)
    pedido.estado_recepcion = 3
    pedido.save()
    messages.warning(request, f'Pedido {pedido.id} anulado')
    return redirect('listarPedidos')

def ActivarPedido(request, id):
    pedido = ORDEN_PEDIDO.objects.get(id = id)
    pedido.estado_recepcion = 0
    pedido.save()
    messages.warning(request, f'Pedido {pedido.id} activado')
    return redirect('listarPedidos')

##******************************************************************

##**************************Categoria proveedor**********************************
@login_required(login_url="login")
def CategoriasProvListar(request):

    categorias = CATEGORIA_PROVEEDOR.objects.all()

    return render(request, 'categoria_proveedor/listar.html', {'categorias':categorias})

@login_required(login_url="login")
def CategoriaProvCrear(request):

    if request.method == 'POST':

        categoria = request.POST.get('descripcion')

        categProv = CATEGORIA_PROVEEDOR.objects.create(
            descripcion = categoria
        )
        categProv.save()
        messages.warning(request, 'Categoria Proveedor creada con exito')
        return redirect('listarCategoriasProv')

    return render(request, 'categoria_proveedor/crear.html')

@login_required(login_url="login")
def CategoriaProvActualizar(request, id):

    categoriaProv = CATEGORIA_PROVEEDOR.objects.get(id=id)
    form = FormCategProv(request.POST or None, instance=categoriaProv)

    if request.method == 'POST':

        categoria = request.POST.get('descripcion')
        categoriaProv.descripcion = categoria
        categoriaProv.save()
        messages.warning(request, 'Categoria Proveedor actualizada con exito')
        return redirect('listarCategoriasProv')

    return render(request, 'categoria_proveedor/actualizar.html',{'form':form})
##******************************************************************

##**************************Boleta***************************************
@login_required(login_url="login")
def BoletaListado(request):
    boletas = BOLETA.objects.all()

    return render(request, "boletas/listar.html", {'boletas':boletas})

@login_required(login_url="login")
def BoletaDetalle(request, id):
    boleta = BOLETA.objects.filter(id=id)
    detalles = DETALLE_BOLETA.objects.filter(boleta=id)
    context = {
        'boleta':boleta,
        'detalles':detalles,
        'id':id
    }
    return render(request, 'boletas/detalles.html', context)

def DesactivarBoleta(request, id):
    boleta = BOLETA.objects.get(id = id)
    boleta.estado = 0
    boleta.save()
    messages.warning(request, f'Boleta {boleta.id} desactivado')
    return redirect('listarBoletas')

def ActivarBoleta(request, id):
    boleta = BOLETA.objects.get(id = id)
    boleta.estado = 1
    boleta.save()
    messages.warning(request, f'Boleta {boleta.id} activado')
    return redirect('listarBoletas')

##********************************************************************

##******************************Informes********************************
@login_required(login_url="login")
def CreacionInformesProductos(request):

    formP = FormProducto()
    formT = FormFamiliaProd()
    formU = FormBoleta()
    formC = FormClientesParaVenta()
    formCliente = FormClientesInforme()
    formProveedor = FormProductoProv()
    formCatProv = FormProveedor()
    formProveeOrden = FormInformeOrdenPedido()
    FormSeg = FormSeguimientoPagina()

    if request.method == 'POST':

        vistaPrevia = request.POST.get('vistaPrevia')
        descargarInforme = request.POST.get('descargarInforme')
 
        productos = request.POST.get('productos')
        precio = request.POST.get('precio')
        descripcion = request.POST.get('descripcion')
        nombre = request.POST.get('nombre')
        precioCompra = request.POST.get('precioCompra')
        stockCritico = request.POST.get('stockCritico')
    
        fechaVencimiento = request.POST.get('fechaVencimiento')
        codigoBarra = request.POST.get('codigoBarra')

        stock = request.POST.get('stock')
        stockCheck = request.POST.get('stockCheck')

        estado = request.POST.get('estado')
        estadoCheck = request.POST.get('estadoCheck')

        familiaProducto = request.POST.get('familiaProducto')
        familiaProductoCheck = request.POST.get('familiaProductoCheck')
        nomFamiliaProducto = request.POST.get('familia_producto')
        nomFamiliaProducto = FAMILIA_PRODUCTO.objects.filter(id = nomFamiliaProducto)
        for fam in nomFamiliaProducto:
            nomFamiliaProducto = fam.descripcion
        
        tipoProducto = request.POST.get('tipoProducto')
        tipoProductoCheck = request.POST.get('tipoProductoCheck')
        nomTipoProducto = request.POST.get('tipo_producto')
        nomTipoProducto = TIPO_PRODUCTO.objects.filter(id = nomTipoProducto)
        for tip in nomTipoProducto:
            nomTipoProducto = tip.descripcion

        visitasPagina = request.POST.get('visitas')

        lista = []
        visitas = []

        tipoInforme = request.POST.get('informeCheck')

        if productos == "on":
            val = PRODUCTO.objects.all().values_list("id","nombre","precio","descripcion","precio_compra","stock", "stock_critico","estado","fecha_vencimiento","codigo_barra","familia_producto__descripcion","familia_producto__tipo_producto__descripcion").order_by("id")
            lista.append(["id","nombre","precio","descripcion","precio_compra","stock", "stock_critico","estado","fecha_vencimiento","codigo_barra","familia_producto","tipo_producto"])

            for valores in val:

                lista.append(list(valores))

            if nombre == None:
                np_array = np.array(lista)
                val = np.where(np_array == "nombre")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)


            if descripcion == None:

                np_array = np.array(lista)
                val = np.where(np_array == "descripcion")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)

            if precio == None:

                np_array = np.array(lista)
                val = np.where(np_array == "precio")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)


            if precioCompra == None:

                np_array = np.array(lista)
                val = np.where(np_array == "precio_compra")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)

            if stockCritico == None:
                np_array = np.array(lista)
                val = np.where(np_array == "stock_critico")
                val = int(val[1])
                
                for valor in lista:

                    valor.pop(val)

            if estado == None:
                np_array = np.array(lista)
                val = np.where(np_array == "estado")
                val = int(val[1])
                
                for valor in lista:

                    valor.pop(val)

            if fechaVencimiento == None:

                np_array = np.array(lista)
                val = np.where(np_array == "fecha_vencimiento")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)

            if codigoBarra == None:

                np_array = np.array(lista)
                val = np.where(np_array == "codigo_barra")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)

            if stock == None:
                np_array = np.array(lista)
                val = np.where(np_array == "stock")
                val = int(val[1])
                for valor in lista:

                    valor.pop(val)

            if familiaProducto == None:

                np_array = np.array(lista)
                val = np.where(np_array == "familia_producto")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)

            if tipoProducto == None:

                np_array = np.array(lista)
                val = np.where(np_array == "tipo_producto")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)

            np_array = []
            np_array = np.array(lista)
            np_arrayProd = np.array(lista)

            if stockCheck == "conStock":
                np_array = np.array(lista)
                val = np.where(np_array == "stock")
                val = int(val[1])
                cont = 0

                for valores in lista:
                    numero = np_arrayProd[cont][val]
                    print(numero)
                    if cont > 0:

                        if int(numero) < 1:
                           
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1
                            
                    cont += 1

            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)
            np_array = np.array(lista)

            if stockCheck == "sinStock":
                np_array = np.array(lista)
                val = np.where(np_array == "stock")
                val = int(val[1])
                cont = 0
                
                for valores in lista:
                    
                    numero = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if int(numero) > 0:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1
                            
            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)

            if estadoCheck == "disponible":
                np_array = np.array(lista)
                val = np.where(np_array == "estado")
                val = int(val[1])
                cont = 0

                for valores in lista:
                    
                    numero = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if int(numero) < 1:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1
                            
            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)

            if estadoCheck == "noDisponible":
                np_array = np.array(lista)
                val = np.where(np_array == "estado")
                val = int(val[1])
                cont = 0

                for valores in lista:
                    
                    numero = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if int(numero) > 0:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1
                            
            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)

            if familiaProductoCheck == "porNombreF":
                np_array = np.array(lista)
                val = np.where(np_array == "familia_producto")
                val = int(val[1])
                cont = 0

                for valores in lista:
                    
                    nombre = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if nombre != nomFamiliaProducto:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1
                            
            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)

            if tipoProductoCheck == "porNombreT":
                np_array = np.array(lista)
                val = np.where(np_array == "tipo_producto")
                val = int(val[1])
                cont = 0

                for valores in lista:
                    
                    nombre = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if nombre != nomTipoProducto:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1
                            
            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)

        if visitasPagina == "on":
            
            paginaVisitada = request.POST.get('paginaVisitada')
            fechaVisitasP = request.POST.get('fechaVisitasP')

            usuarioVisitasPagina = request.POST.get('usuarioPaginaVisitada')
            usuarioVisitas = request.POST.get('usuarioVisitasPaginasCheck')
            nomUsuario = request.POST.get('usuario')
            nomUsuario = User.objects.filter(id = nomUsuario)
            for us in nomUsuario:
                nomUsuario = us.username

            val = SEGUIMIENTO_PAGINA.objects.all().values_list("id","pagina_visitada","fecha_ingreso","usuario__username").order_by("id")
            visitas.append(["id","pagina_visitada","fecha_ingreso","usuario"])
            
            for valores in val:

                visitas.append(list(valores))

            if paginaVisitada == None:
                np_array = np.array(visitas)
                val = np.where(np_array == "pagina_visitada")
                val = int(val[1])

                for valor in visitas:

                    valor.pop(val)

            if fechaVisitasP == None:
                np_array = np.array(visitas)
                val = np.where(np_array == "fecha_ingreso")
                val = int(val[1])

                for valor in visitas:

                    valor.pop(val)

            if usuarioVisitasPagina == None:
                np_array = np.array(visitas)
                val = np.where(np_array == "usuario")
                val = int(val[1])

                for valor in visitas:

                    valor.pop(val)

            np_array = []
            np_array = np.array(visitas)
            np_arrayProd = np.array(visitas)
            
            if usuarioVisitas == "porNombreVisitasP":
                np_array = np.array(visitas)
                val = np.where(np_array == "usuario")
                val = int(val[1])
                cont = 0

                for valores in visitas:
                    
                    nombre = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if nombre != nomUsuario:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1

            visitas = []
            np_array = []
            visitas = np_arrayProd.tolist()
            np_arrayProd = np.array(visitas)

        if tipoInforme == "informeExcel":
            
            nombre_archivo = "Productos"
            if visitas == []:
                return  creacion_excel(nombre_archivo, lista)
            else:
                return  creacion_excel(nombre_archivo, lista, visitas)
            

        if tipoInforme == "informePdf":

            tipo_doc = 'pdf'
            extension = 'pdf'
            
            nombre = 'Productos'
            
            if vistaPrevia:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=False)
                else:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=False, visitas=visitas)
            else:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=True)
                else:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=True, visitas=visitas)
 
        if tipoInforme == "informeWord": 

            tipo_doc = 'ms-word'
            extension = 'docx'
            
            nombre = 'Productos'
            if vistaPrevia:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=False)
                else:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=False, visitas=visitas)
            else:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=True)
                else:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=True, visitas=visitas)

    context = {
        'formP':formP,
        'formT':formT,
        'formU':formU,
        'formC':formC,
        'formCliente':formCliente,
        'formProveedor':formProveedor,
        'formCatProv':formCatProv,
        'formProveeOrden':formProveeOrden,
        
        'FormSeg':FormSeg,

    }
    return render(request, 'informes/informe_productos.html', context)

@login_required(login_url="login")
def CreacionInformesBoletas(request):

    formP = FormProducto()
    formT = FormFamiliaProd()
    formU = FormBoleta()
    formC = FormClientesParaVenta()
    formCliente = FormClientesInforme()
    formProveedor = FormProductoProv()
    formCatProv = FormProveedor()
    formProveeOrden = FormInformeOrdenPedido()
    FormSeg = FormSeguimientoPagina()

    if request.method == 'POST':

        boletas = request.POST.get('boletas')
        fecha = request.POST.get('fecha')
        total = request.POST.get('total')

        usuarioB = request.POST.get('usuarioB')
        usuarioBoletaCheck = request.POST.get('usuarioBoletaCheck')
        nomUsuario = request.POST.get('usuario')
        nomUsuario = User.objects.filter(id = nomUsuario)
        for us in nomUsuario:
            nomUsuario = us.username

        clienteB = request.POST.get('clienteB')
        clienteBoletaCheck = request.POST.get('clienteBoletaCheck')
        nomCliente = request.POST.get('cliente')
        nomCliente = CLIENTE.objects.filter(id = nomCliente)
        for cli in nomCliente:
            nomCliente = cli.nombre

        estadoB = request.POST.get('estadoB')
        estadoBoletaCheck = request.POST.get('estadoBoletaCheck')

        detalleBoleta = request.POST.get('detalleBoleta')
        productoDetalle = request.POST.get('productoDetalle')
        cantidadDetalle = request.POST.get('cantidadDetalle')
        precioDetalle = request.POST.get('precioDetalle')

        tipoInforme = request.POST.get('informeCheck')

        vistaPrevia = request.POST.get('vistaPrevia')
        descargarInforme = request.POST.get('descargarInforme')
        visitasPagina = request.POST.get('visitas')

        lista = []
        visitas = []

        if boletas == "on":

            val = DETALLE_BOLETA.objects.all().values_list("boleta__id","boleta__fecha_boleta","boleta__total_a_pagar","boleta__usuario__username","boleta__cliente__nombre","boleta__estado","producto__nombre","cantidad","monto_a_pagar").order_by("boleta__id")
            lista.append(["id","fecha","total","usuario","cliente","estado","producto","cantidad","precio"])

            for valores in val:

                lista.append(list(valores))

            if fecha == None:
                np_array = np.array(lista)
                val = np.where(np_array == "fecha")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)


            if total == None:

                np_array = np.array(lista)
                val = np.where(np_array == "total")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)

            if usuarioB == None:

                np_array = np.array(lista)
                val = np.where(np_array == "usuario")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)


            if clienteB == None:

                np_array = np.array(lista)
                val = np.where(np_array == "cliente")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)

            if estadoB == None:
                np_array = np.array(lista)
                val = np.where(np_array == "estado")
                val = int(val[1])
                
                for valor in lista:

                    valor.pop(val)

            if detalleBoleta == None:
                
                np_array = np.array(lista)
                val = np.where(np_array == "producto")
                val = int(val[1])
                
                for valor in lista:

                    valor.pop(val)

            if detalleBoleta == None:

                np_array = np.array(lista)
                val = np.where(np_array == "cantidad")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)

            if detalleBoleta == None:

                np_array = np.array(lista)
                val = np.where(np_array == "precio")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)

            np_array = []
            np_array = np.array(lista)
            np_arrayProd = np.array(lista)

            if usuarioBoletaCheck == "porNombreB":
                np_array = np.array(lista)
                val = np.where(np_array == "usuario")
                val = int(val[1])
                cont = 0

                for valores in lista:
                    
                    nombre = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if nombre != nomUsuario:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1

            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)

            if clienteBoletaCheck == "porNombreB":
                np_array = np.array(lista)
                val = np.where(np_array == "cliente")
                val = int(val[1])
                cont = 0

                for valores in lista:
                    
                    nombre = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if nombre != nomCliente:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1
                            
            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)
            
            if estadoBoletaCheck == "activasB":
                np_array = np.array(lista)
                val = np.where(np_array == "estado")
                val = int(val[1])
                cont = 0

                for valores in lista:
                    
                    numero = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if int(numero) < 1:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1
                            
            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)

            if estadoBoletaCheck == "anuladasB":
                np_array = np.array(lista)
                val = np.where(np_array == "estado")
                val = int(val[1])
                cont = 0

                for valores in lista:
                    
                    numero = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if int(numero) > 0:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1
                            
            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)

        if visitasPagina == "on":
            
            paginaVisitada = request.POST.get('paginaVisitada')
            fechaVisitasP = request.POST.get('fechaVisitasP')

            usuarioVisitasPagina = request.POST.get('usuarioPaginaVisitada')
            usuarioVisitas = request.POST.get('usuarioVisitasPaginasCheck')
            nomUsuario = request.POST.get('usuario')
            nomUsuario = User.objects.filter(id = nomUsuario)
            for us in nomUsuario:
                nomUsuario = us.username

            val = SEGUIMIENTO_PAGINA.objects.all().values_list("id","pagina_visitada","fecha_ingreso","usuario__username").order_by("id")
            visitas.append(["id","pagina_visitada","fecha_ingreso","usuario"])
            
            for valores in val:

                visitas.append(list(valores))

            if paginaVisitada == None:
                np_array = np.array(visitas)
                val = np.where(np_array == "pagina_visitada")
                val = int(val[1])

                for valor in visitas:

                    valor.pop(val)

            if fechaVisitasP == None:
                np_array = np.array(visitas)
                val = np.where(np_array == "fecha_ingreso")
                val = int(val[1])

                for valor in visitas:

                    valor.pop(val)

            if usuarioVisitasPagina == None:
                np_array = np.array(visitas)
                val = np.where(np_array == "usuario")
                val = int(val[1])

                for valor in visitas:

                    valor.pop(val)

            np_array = []
            np_array = np.array(visitas)
            np_arrayProd = np.array(visitas)
            
            if usuarioVisitas == "porNombreVisitasP":
                np_array = np.array(visitas)
                val = np.where(np_array == "usuario")
                val = int(val[1])
                cont = 0

                for valores in visitas:
                    
                    nombre = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if nombre != nomUsuario:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1

            visitas = []
            np_array = []
            visitas = np_arrayProd.tolist()
            np_arrayProd = np.array(visitas)
            
        if tipoInforme == "informeExcel":

            nombre_archivo = "Boletas"
            if visitas == []:
                return  creacion_excel(nombre_archivo, lista)
            else:
                return  creacion_excel(nombre_archivo, lista, visitas)

        if tipoInforme == "informePdf":
            
            tipo_doc = 'pdf'
            extension = 'pdf'
            
            nombre = 'Boletas'
            if vistaPrevia:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=False)
                else:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=False, visitas=visitas)
            else:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=True)
                else:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=True, visitas=visitas)

        if tipoInforme == "informeWord": 

            tipo_doc = 'ms-word'
            extension = 'docx'
            
            nombre = 'Boletas'

            if vistaPrevia:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=False)
                else:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=False, visitas=visitas)
            else:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=True)
                else:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=True, visitas=visitas)

    context = {
        'formP':formP,
        'formT':formT,
        'formU':formU,
        'formC':formC,
        'formCliente':formCliente,
        'formProveedor':formProveedor,
        'formCatProv':formCatProv,
        'formProveeOrden':formProveeOrden,
        'FormSeg':FormSeg,

    }
    return render(request, 'informes/informe_boletas.html', context)

@login_required(login_url="login")
def CreacionInformesClientes(request):

    formP = FormProducto()
    formT = FormFamiliaProd()
    formU = FormBoleta()
    formC = FormClientesParaVenta()
    formCliente = FormClientesInforme()
    formProveedor = FormProductoProv()
    formCatProv = FormProveedor()
    formProveeOrden = FormInformeOrdenPedido()
    FormSeg = FormSeguimientoPagina()

    if request.method == 'POST':

        clientes = request.POST.get('clientes')
        runC = request.POST.get('runC')
        telefonoC = request.POST.get('telefonoC')
        correoC = request.POST.get('correoC')
        direccionC = request.POST.get('direccionC')
        
        nombreC = request.POST.get('nombreC')
        nombreClienteCheck = request.POST.get('nombreClienteCheck')
        nomCliente = request.POST.get('cliente')
        nomCliente = CLIENTE.objects.filter(id = nomCliente)
        for us in nomCliente:
            nomCliente = us.nombre

        estadoIC = request.POST.get('estadoIC')
        estadoBoletaCheck = request.POST.get('estadoIClienteCheck')

        tipoInforme = request.POST.get('informeCheck')

        vistaPrevia = request.POST.get('vistaPrevia')
        descargarInforme = request.POST.get('descargarInforme')
        visitasPagina = request.POST.get('visitas')

        lista = []
        visitas = []

        if clientes == "on":

            val = CLIENTE.objects.all().values_list("id","run","nombre","telefono","correo","direccion","estado").order_by("id")
            lista.append(["id","run","nombre","telefono","correo","direccion","estado"])

            for valores in val:

                lista.append(list(valores))

            if runC == None:
                np_array = np.array(lista)
                val = np.where(np_array == "run")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)


            if nombreC == None:

                np_array = np.array(lista)
                val = np.where(np_array == "nombre")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)

            if telefonoC == None:

                np_array = np.array(lista)
                val = np.where(np_array == "telefono")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)


            if correoC == None:

                np_array = np.array(lista)
                val = np.where(np_array == "correo")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)

            if direccionC == None:
                np_array = np.array(lista)
                val = np.where(np_array == "direccion")
                val = int(val[1])
                
                for valor in lista:

                    valor.pop(val)

            if estadoIC == None:
                
                np_array = np.array(lista)
                val = np.where(np_array == "estado")
                val = int(val[1])
                
                for valor in lista:

                    valor.pop(val)

            np_array = []
            np_array = np.array(lista)
            np_arrayProd = np.array(lista)

            if nombreClienteCheck == "porNombreC":
                np_array = np.array(lista)
                val = np.where(np_array == "nombre")
                val = int(val[1])
                cont = 0

                for valores in lista:
                    
                    nombre = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if nombre != nomCliente:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1

            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)
            
            if estadoBoletaCheck == "activoIC":
                np_array = np.array(lista)
                val = np.where(np_array == "estado")
                val = int(val[1])
                cont = 0

                for valores in lista:
                    
                    numero = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if int(numero) < 1:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1
                            
            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)

            if estadoBoletaCheck == "bloqueadoIC":
                np_array = np.array(lista)
                val = np.where(np_array == "estado")
                val = int(val[1])
                cont = 0

                for valores in lista:
                    
                    numero = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if int(numero) > 0:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1
                            
            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)

        if visitasPagina == "on":
            
            paginaVisitada = request.POST.get('paginaVisitada')
            fechaVisitasP = request.POST.get('fechaVisitasP')

            usuarioVisitasPagina = request.POST.get('usuarioPaginaVisitada')
            usuarioVisitas = request.POST.get('usuarioVisitasPaginasCheck')
            nomUsuario = request.POST.get('usuario')
            nomUsuario = User.objects.filter(id = nomUsuario)
            for us in nomUsuario:
                nomUsuario = us.username

            val = SEGUIMIENTO_PAGINA.objects.all().values_list("id","pagina_visitada","fecha_ingreso","usuario__username").order_by("id")
            visitas.append(["id","pagina_visitada","fecha_ingreso","usuario"])
            
            for valores in val:

                visitas.append(list(valores))

            if paginaVisitada == None:
                np_array = np.array(visitas)
                val = np.where(np_array == "pagina_visitada")
                val = int(val[1])

                for valor in visitas:

                    valor.pop(val)

            if fechaVisitasP == None:
                np_array = np.array(visitas)
                val = np.where(np_array == "fecha_ingreso")
                val = int(val[1])

                for valor in visitas:

                    valor.pop(val)

            if usuarioVisitasPagina == None:
                np_array = np.array(visitas)
                val = np.where(np_array == "usuario")
                val = int(val[1])

                for valor in visitas:

                    valor.pop(val)

            np_array = []
            np_array = np.array(visitas)
            np_arrayProd = np.array(visitas)
            
            if usuarioVisitas == "porNombreVisitasP":
                np_array = np.array(visitas)
                val = np.where(np_array == "usuario")
                val = int(val[1])
                cont = 0

                for valores in visitas:
                    
                    nombre = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if nombre != nomUsuario:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1

            visitas = []
            np_array = []
            visitas = np_arrayProd.tolist()
            np_arrayProd = np.array(visitas)
            
        if tipoInforme == "informeExcel":

            nombre_archivo = "Clientes"
            if visitas == []:
                return  creacion_excel(nombre_archivo, lista)
            else:
                return  creacion_excel(nombre_archivo, lista, visitas)

        if tipoInforme == "informePdf":
            
            tipo_doc = 'pdf'
            extension = 'pdf'
            
            nombre = 'Clientes'
            if vistaPrevia:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=False)
                else:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=False, visitas=visitas)
            else:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=True)
                else:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=True, visitas=visitas)

        if tipoInforme == "informeWord": 

            tipo_doc = 'ms-word'
            extension = 'docx'
            
            nombre = 'Clientes'

            if vistaPrevia:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=False)
                else:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=False, visitas=visitas)
            else:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=True)
                else:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=True, visitas=visitas)

    context = {
        'formP':formP,
        'formT':formT,
        'formU':formU,
        'formC':formC,
        'formCliente':formCliente,
        'formProveedor':formProveedor,
        'formCatProv':formCatProv,
        'formProveeOrden':formProveeOrden,
        'FormSeg':FormSeg,

    }
    return render(request, 'informes/informe_clientes.html', context)

@login_required(login_url="login")
def CreacionInformesFiados(request):

    formP = FormProducto()
    formT = FormFamiliaProd()
    formU = FormBoleta()
    formC = FormClientesParaVenta()
    formCliente = FormClientesInforme()
    formProveedor = FormProductoProv()
    formCatProv = FormProveedor()
    formProveeOrden = FormInformeOrdenPedido()
    FormSeg = FormSeguimientoPagina()

    if request.method == 'POST':

        fiados = request.POST.get('fiados')
        montoF = request.POST.get('montoF')
        fechaF = request.POST.get('fechaF')
        fechaFinalF = request.POST.get('fechaFinalF')
        clienteFiado = request.POST.get('clienteFiado')

        estadoF = request.POST.get('estadoF')
        estadoFiadoCheck = request.POST.get('estadoFiadoCheck')

        detallePagosF = request.POST.get('detallePagosF')
        montoAbonadoF = request.POST.get('montoAbonadoF')
        fechaAbonoF = request.POST.get('fechaAbonoF')
        

        tipoInforme = request.POST.get('informeCheck')

        vistaPrevia = request.POST.get('vistaPrevia')
        descargarInforme = request.POST.get('descargarInforme')
        visitasPagina = request.POST.get('visitas')

        lista = []
        visitas = []

        if fiados == "on":

            val = DETALLE_FIADO.objects.all().values_list("id","pago_fiado__cliente__nombre","pago_fiado__monto","pago_fiado__fecha_creacion","pago_fiado__fecha_final","pago_fiado__estado","monto_abonado","fecha_abono").order_by("id")
            lista.append(["id","cliente","monto","fecha_creacion","fecha_final","estado","monto_abonado","fecha_abono"])

            for valores in val:

                lista.append(list(valores))

            if montoF == None:
                np_array = np.array(lista)
                val = np.where(np_array == "monto")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)


            if fechaF == None:

                np_array = np.array(lista)
                val = np.where(np_array == "fecha_creacion")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)

            if fechaFinalF == None:

                np_array = np.array(lista)
                val = np.where(np_array == "razon_social")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)


            if clienteFiado == None:

                np_array = np.array(lista)
                val = np.where(np_array == "cliente")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)

            if estadoF == None:
                np_array = np.array(lista)
                val = np.where(np_array == "estado")
                val = int(val[1])
                
                for valor in lista:

                    valor.pop(val)

            if montoAbonadoF == None:
                
                np_array = np.array(lista)
                val = np.where(np_array == "monto_abonado")
                val = int(val[1])
                
                for valor in lista:

                    valor.pop(val)

            if fechaAbonoF == None:
                
                np_array = np.array(lista)
                val = np.where(np_array == "fecha_abono")
                val = int(val[1])
                
                for valor in lista:

                    valor.pop(val)

            np_array = []
            np_array = np.array(lista)
            np_arrayProd = np.array(lista)
            
            if estadoFiadoCheck == "alDia":
                np_array = np.array(lista)
                val = np.where(np_array == "estado")
                val = int(val[1])
                cont = 0

                for valores in lista:
                    
                    numero = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if int(numero) != 1:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1
                            
            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)

            if estadoFiadoCheck == "atrasados":
                np_array = np.array(lista)
                val = np.where(np_array == "estado")
                val = int(val[1])
                cont = 0

                for valores in lista:
                    
                    numero = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if int(numero) != 2:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1
                            
            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)

        if visitasPagina == "on":
            
            paginaVisitada = request.POST.get('paginaVisitada')
            fechaVisitasP = request.POST.get('fechaVisitasP')

            usuarioVisitasPagina = request.POST.get('usuarioPaginaVisitada')
            usuarioVisitas = request.POST.get('usuarioVisitasPaginasCheck')
            nomUsuario = request.POST.get('usuario')
            nomUsuario = User.objects.filter(id = nomUsuario)
            for us in nomUsuario:
                nomUsuario = us.username

            val = SEGUIMIENTO_PAGINA.objects.all().values_list("id","pagina_visitada","fecha_ingreso","usuario__username").order_by("id")
            visitas.append(["id","pagina_visitada","fecha_ingreso","usuario"])
            
            for valores in val:

                visitas.append(list(valores))

            if paginaVisitada == None:
                np_array = np.array(visitas)
                val = np.where(np_array == "pagina_visitada")
                val = int(val[1])

                for valor in visitas:

                    valor.pop(val)

            if fechaVisitasP == None:
                np_array = np.array(visitas)
                val = np.where(np_array == "fecha_ingreso")
                val = int(val[1])

                for valor in visitas:

                    valor.pop(val)

            if usuarioVisitasPagina == None:
                np_array = np.array(visitas)
                val = np.where(np_array == "usuario")
                val = int(val[1])

                for valor in visitas:

                    valor.pop(val)

            np_array = []
            np_array = np.array(visitas)
            np_arrayProd = np.array(visitas)
            
            if usuarioVisitas == "porNombreVisitasP":
                np_array = np.array(visitas)
                val = np.where(np_array == "usuario")
                val = int(val[1])
                cont = 0

                for valores in visitas:
                    
                    nombre = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if nombre != nomUsuario:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1

            visitas = []
            np_array = []
            visitas = np_arrayProd.tolist()
            np_arrayProd = np.array(visitas)
            
        if tipoInforme == "informeExcel":

            nombre_archivo = "Fiados"
            if visitas == []:
                return  creacion_excel(nombre_archivo, lista)
            else:
                return  creacion_excel(nombre_archivo, lista, visitas)

        if tipoInforme == "informePdf":
            
            tipo_doc = 'pdf'
            extension = 'pdf'
            
            nombre = 'Fiados'
            if vistaPrevia:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=False)
                else:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=False, visitas=visitas)
            else:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=True)
                else:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=True, visitas=visitas)

        if tipoInforme == "informeWord": 

            tipo_doc = 'ms-word'
            extension = 'docx'
            
            nombre = 'Fiados'

            if vistaPrevia:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=False)
                else:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=False, visitas=visitas)
            else:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=True)
                else:
                    return creacion_doc(lista,tipo_doc,A3,nombre,extension, valor=True, visitas=visitas)

    context = {
        'formP':formP,
        'formT':formT,
        'formU':formU,
        'formC':formC,
        'formCliente':formCliente,
        'formProveedor':formProveedor,
        'formCatProv':formCatProv,
        'formProveeOrden':formProveeOrden,
        'FormSeg':FormSeg,

    }
    return render(request, 'informes/informe_fiados.html', context)

@login_required(login_url="login")
def CreacionInformesProveedores(request):

    formP = FormProducto()
    formT = FormFamiliaProd()
    formU = FormBoleta()
    formC = FormClientesParaVenta()
    formCliente = FormClientesInforme()
    formProveedor = FormProductoProv()
    formCatProv = FormProveedor()
    formProveeOrden = FormInformeOrdenPedido()
    FormSeg = FormSeguimientoPagina()

    if request.method == 'POST':

        proveedores = request.POST.get('proveedores')
        correoP = request.POST.get('correoP')
        telefonoP = request.POST.get('telefonoP')
        direccionP = request.POST.get('direccionP')
        
        estadoP = request.POST.get('estadoP')
        estadoProveedorCheck = request.POST.get('estadoProveedorCheck')

        razonSocialP = request.POST.get('razonSocialP')
        razonSocialPCheck = request.POST.get('razonSocialPCheck')
        nomProveedor = request.POST.get('proveedor')
        nomProveedor = PROVEEDOR.objects.filter(id = nomProveedor)
        for us in nomProveedor:
            nomProveedor = us.razon_social

        categoriaP = request.POST.get('categoriaP')
        categoriaProvCheck = request.POST.get('categoriaProveedorCheck')
        nomCategoria = request.POST.get('categoria_proveedor')
        nomCategoria = CATEGORIA_PROVEEDOR.objects.filter(id = nomCategoria)
        for us in nomCategoria:
            nomCategoria = us.descripcion

        tipoInforme = request.POST.get('informeCheck')

        vistaPrevia = request.POST.get('vistaPrevia')
        descargarInforme = request.POST.get('descargarInforme')
        visitasPagina = request.POST.get('visitas')

        lista = []
        visitas = []

        if proveedores == "on":

            val = PROVEEDOR.objects.all().values_list("id","razon_social","correo","telefono","direccion","estado","categoria_proveedor__descripcion").order_by("id")
            lista.append(["id","razon_social","correo","telefono","direccion","estado","categoria"])

            for valores in val:

                lista.append(list(valores))

            if correoP == None:
                np_array = np.array(lista)
                val = np.where(np_array == "correo")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)


            if telefonoP == None:

                np_array = np.array(lista)
                val = np.where(np_array == "telefono")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)

            if razonSocialP == None:

                np_array = np.array(lista)
                val = np.where(np_array == "razon_social")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)


            if direccionP == None:

                np_array = np.array(lista)
                val = np.where(np_array == "direccion")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)

            if estadoP == None:
                np_array = np.array(lista)
                val = np.where(np_array == "estado")
                val = int(val[1])
                
                for valor in lista:

                    valor.pop(val)

            if categoriaP == None:
                
                np_array = np.array(lista)
                val = np.where(np_array == "categoria")
                val = int(val[1])
                
                for valor in lista:

                    valor.pop(val)

            np_array = []
            np_array = np.array(lista)
            np_arrayProd = np.array(lista)

            if razonSocialPCheck == "porRazonSocialP":
                np_array = np.array(lista)
                val = np.where(np_array == "razon_social")
                val = int(val[1])
                cont = 0

                for valores in lista:
                    
                    nombre = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if nombre != nomProveedor:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1

            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)

            if categoriaProvCheck == "porCategoriaP":
                np_array = np.array(lista)
                val = np.where(np_array == "categoria")
                val = int(val[1])
                cont = 0

                for valores in lista:
                    
                    nombre = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if nombre != nomCategoria:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1

            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)
            
            if estadoProveedorCheck == "activosP":
                np_array = np.array(lista)
                val = np.where(np_array == "estado")
                val = int(val[1])
                cont = 0

                for valores in lista:
                    
                    numero = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if int(numero) < 1:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1
                            
            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)

            if estadoProveedorCheck == "bloqueadosP":
                np_array = np.array(lista)
                val = np.where(np_array == "estado")
                val = int(val[1])
                cont = 0

                for valores in lista:
                    
                    numero = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if int(numero) > 0:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1
                            
            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)

        if visitasPagina == "on":
            
            paginaVisitada = request.POST.get('paginaVisitada')
            fechaVisitasP = request.POST.get('fechaVisitasP')

            usuarioVisitasPagina = request.POST.get('usuarioPaginaVisitada')
            usuarioVisitas = request.POST.get('usuarioVisitasPaginasCheck')
            nomUsuario = request.POST.get('usuario')
            nomUsuario = User.objects.filter(id = nomUsuario)
            for us in nomUsuario:
                nomUsuario = us.username

            val = SEGUIMIENTO_PAGINA.objects.all().values_list("id","pagina_visitada","fecha_ingreso","usuario__username").order_by("id")
            visitas.append(["id","pagina_visitada","fecha_ingreso","usuario"])
            
            for valores in val:

                visitas.append(list(valores))

            if paginaVisitada == None:
                np_array = np.array(visitas)
                val = np.where(np_array == "pagina_visitada")
                val = int(val[1])

                for valor in visitas:

                    valor.pop(val)

            if fechaVisitasP == None:
                np_array = np.array(visitas)
                val = np.where(np_array == "fecha_ingreso")
                val = int(val[1])

                for valor in visitas:

                    valor.pop(val)

            if usuarioVisitasPagina == None:
                np_array = np.array(visitas)
                val = np.where(np_array == "usuario")
                val = int(val[1])

                for valor in visitas:

                    valor.pop(val)

            np_array = []
            np_array = np.array(visitas)
            np_arrayProd = np.array(visitas)
            
            if usuarioVisitas == "porNombreVisitasP":
                np_array = np.array(visitas)
                val = np.where(np_array == "usuario")
                val = int(val[1])
                cont = 0

                for valores in visitas:
                    
                    nombre = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if nombre != nomUsuario:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1

            visitas = []
            np_array = []
            visitas = np_arrayProd.tolist()
            np_arrayProd = np.array(visitas)
            
        if tipoInforme == "informeExcel":

            nombre_archivo = "Proveedores"
            if visitas == []:
                return  creacion_excel(nombre_archivo, lista)
            else:
                return  creacion_excel(nombre_archivo, lista, visitas)

        if tipoInforme == "informePdf":
            
            tipo_doc = 'pdf'
            extension = 'pdf'
            
            nombre = 'Proveedores'
            if vistaPrevia:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=False)
                else:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=False, visitas=visitas)
            else:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=True)
                else:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=True, visitas=visitas)

        if tipoInforme == "informeWord": 

            tipo_doc = 'ms-word'
            extension = 'docx'
            
            nombre = 'Proveedores'

            if vistaPrevia:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=False)
                else:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=False, visitas=visitas)
            else:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=True)
                else:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=True, visitas=visitas)

    context = {
        'formP':formP,
        'formT':formT,
        'formU':formU,
        'formC':formC,
        'formCliente':formCliente,
        'formProveedor':formProveedor,
        'formCatProv':formCatProv,
        'formProveeOrden':formProveeOrden,
        'FormSeg':FormSeg,

    }
    return render(request, 'informes/informe_proveedores.html',context)

@login_required(login_url="login")
def CreacionInformesPedidos(request):

    formP = FormProducto()
    formT = FormFamiliaProd()
    formU = FormBoleta()
    formC = FormClientesParaVenta()
    formCliente = FormClientesInforme()
    formProveedor = FormProductoProv()
    formCatProv = FormProveedor()
    formProveeOrden = FormInformeOrdenPedido()
    FormSeg = FormSeguimientoPagina()

    if request.method == 'POST':

        ordenes = request.POST.get('ordenes')
        fechaOrdenP = request.POST.get('fechaOrdenP')
        fechaLlegadaOrdenP = request.POST.get('fechaLlegadaOrdenP')
        fechaRecepOrdenP = request.POST.get('fechaRecepOrdenP')
        horaRecepOrdenP = request.POST.get('horaRecepOrdenP')
        cantidadOrdenP = request.POST.get('cantidadOrdenP')
        productoOrdenP = request.POST.get('productoOrdenP')

        estadoRecepcionO = request.POST.get('estadoRecepcionO')
        estadoOrdenPCheck = request.POST.get('estadoOrdenPCheck')

        razonSocialOrdenP = request.POST.get('razonSocialOrdenP')
        razonSocialOrdenPCheck = request.POST.get('razonSocialOrdenPCheck')
        nomProveedor = request.POST.get('proveedor')
        nomProveedor = PROVEEDOR.objects.filter(id = nomProveedor)
        for us in nomProveedor:
            nomProveedor = us.razon_social
        
        tipoInforme = request.POST.get('informeCheck')

        vistaPrevia = request.POST.get('vistaPrevia')
        descargarInforme = request.POST.get('descargarInforme')
        visitasPagina = request.POST.get('visitas')

        lista = []
        visitas = []

        if ordenes == "on":

            val = DETALLE_ORDEN.objects.all().values_list("id","orden_pedido__proveedor__razon_social","orden_pedido__fecha_pedido","orden_pedido__fecha_llegada","orden_pedido__fecha_recepcion","orden_pedido__hora_recepcion","orden_pedido__estado_recepcion","producto__nombre","cantidad").order_by("id")
            lista.append(["id","proveedor","fecha_pedido","fecha_llegada","fecha_recepcion","hora_recepcion","estado","producto","cantidad"])

            for valores in val:

                lista.append(list(valores))

            if fechaOrdenP == None:
                np_array = np.array(lista)
                val = np.where(np_array == "fecha_pedido")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)

            if fechaLlegadaOrdenP == None:

                np_array = np.array(lista)
                val = np.where(np_array == "fecha_llegada")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)

            if fechaRecepOrdenP == None:

                np_array = np.array(lista)
                val = np.where(np_array == "fecha_recepcion")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)


            if horaRecepOrdenP == None:

                np_array = np.array(lista)
                val = np.where(np_array == "hora_recepcion")
                val = int(val[1])

                for valor in lista:

                    valor.pop(val)

            if cantidadOrdenP == None:
                np_array = np.array(lista)
                val = np.where(np_array == "cantidad")
                val = int(val[1])
                
                for valor in lista:

                    valor.pop(val)

            if productoOrdenP == None:
                
                np_array = np.array(lista)
                val = np.where(np_array == "producto")
                val = int(val[1])
                
                for valor in lista:

                    valor.pop(val)

            if estadoRecepcionO == None:
                
                np_array = np.array(lista)
                val = np.where(np_array == "estado")
                val = int(val[1])
                
                for valor in lista:

                    valor.pop(val)

            if razonSocialOrdenP == None:
                
                np_array = np.array(lista)
                val = np.where(np_array == "proveedor")
                val = int(val[1])
                
                for valor in lista:

                    valor.pop(val)
        
            np_array = []
            np_array = np.array(lista)
            np_arrayProd = np.array(lista)

            if razonSocialOrdenPCheck == "porRazonSocialOrdenP":
                np_array = np.array(lista)
                val = np.where(np_array == "proveedor")
                val = int(val[1])
                cont = 0
                print(val)
                for valores in lista:
                    
                    nombre = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if nombre != nomProveedor:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1

            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)

            if estadoOrdenPCheck == "enEsperaOrden":
                np_array = np.array(lista)
                val = np.where(np_array == "estado")
                val = int(val[1])
                cont = 0

                for valores in lista:
                    
                    numero = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if int(numero) > 0:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1
                            
            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)

            if estadoOrdenPCheck == "recepcionadoOrden":
                np_array = np.array(lista)
                val = np.where(np_array == "estado")
                val = int(val[1])
                cont = 0

                for valores in lista:
                    
                    numero = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if int(numero) != 1:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1
                            
            lista = []
            np_array = []
            lista = np_arrayProd.tolist()
            np_arrayProd = np.array(lista)

        if visitasPagina == "on":
            
            paginaVisitada = request.POST.get('paginaVisitada')
            fechaVisitasP = request.POST.get('fechaVisitasP')

            usuarioVisitasPagina = request.POST.get('usuarioPaginaVisitada')
            usuarioVisitas = request.POST.get('usuarioVisitasPaginasCheck')
            nomUsuario = request.POST.get('usuario')
            nomUsuario = User.objects.filter(id = nomUsuario)
            for us in nomUsuario:
                nomUsuario = us.username

            val = SEGUIMIENTO_PAGINA.objects.all().values_list("id","pagina_visitada","fecha_ingreso","usuario__username").order_by("id")
            visitas.append(["id","pagina_visitada","fecha_ingreso","usuario"])
            
            for valores in val:

                visitas.append(list(valores))

            if paginaVisitada == None:
                np_array = np.array(visitas)
                val = np.where(np_array == "pagina_visitada")
                val = int(val[1])

                for valor in visitas:

                    valor.pop(val)

            if fechaVisitasP == None:
                np_array = np.array(visitas)
                val = np.where(np_array == "fecha_ingreso")
                val = int(val[1])

                for valor in visitas:

                    valor.pop(val)

            if usuarioVisitasPagina == None:
                np_array = np.array(visitas)
                val = np.where(np_array == "usuario")
                val = int(val[1])

                for valor in visitas:

                    valor.pop(val)

            np_array = []
            np_array = np.array(visitas)
            np_arrayProd = np.array(visitas)
            
            if usuarioVisitas == "porNombreVisitasP":
                np_array = np.array(visitas)
                val = np.where(np_array == "usuario")
                val = int(val[1])
                cont = 0

                for valores in visitas:
                    
                    nombre = np_arrayProd[cont][val]
                    
                    if cont > 0:

                        if nombre != nomUsuario:
                            
                            np_arrayProd = np.delete(np_arrayProd, cont, axis=0)
                            cont = cont - 1 
                            
                    cont += 1

            visitas = []
            np_array = []
            visitas = np_arrayProd.tolist()
            np_arrayProd = np.array(visitas)
            
        if tipoInforme == "informeExcel":

            nombre_archivo = "Pedidos"
            if visitas == []:
                return  creacion_excel(nombre_archivo, lista)
            else:
                return  creacion_excel(nombre_archivo, lista, visitas)

        if tipoInforme == "informePdf":
            
            tipo_doc = 'pdf'
            extension = 'pdf'
            
            nombre = 'Pedidos'
            if vistaPrevia:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=False)
                else:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=False, visitas=visitas)
            else:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=True)
                else:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=True, visitas=visitas)

        if tipoInforme == "informeWord": 

            tipo_doc = 'ms-word'
            extension = 'docx'
            
            nombre = 'Pedidos'

            if vistaPrevia:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=False)
                else:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=False, visitas=visitas)
            else:
                if visitas == []:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=True)
                else:
                    return creacion_doc(lista,tipo_doc,A2,nombre,extension, valor=True, visitas=visitas)

    context = {
        'formP':formP,
        'formT':formT,
        'formU':formU,
        'formC':formC,
        'formCliente':formCliente,
        'formProveedor':formProveedor,
        'formCatProv':formCatProv,
        'formProveeOrden':formProveeOrden,
        'FormSeg':FormSeg,

    }
    return render(request, 'informes/informe_pedidos.html', context)

def creacion_excel(nombre_archivo, lista, visitas = None):
    book = openpyxl.Workbook()  # Se crea un libro
    sheet = book.active  # Se activa la primera hojar
    sheet.title = f"{nombre_archivo}"
    cont = 0
    cont2 = 0

    for i in lista:
        cont += 1
        for j in i:
            
            cont2 += 1
            val = sheet.cell(row=cont, column=cont2)
            val.value = j
        cont2 = 0    

    cont = 0
    cont2 = 0
    
    if visitas:
        sheet2 = book.create_sheet("Visitas")
        for i in visitas:
            cont += 1
            for j in i:
                
                cont2 += 1
                val = sheet2.cell(row=cont, column=cont2)
                val.value = j
            cont2 = 0 
    # nombre_archivo = "boletas.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename = {0}.xlsx".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    book.save(response)

    return response

def creacion_doc(lista,tipo_doc,tamaño_pagina, nombre, extension, visitas = None, valor = None):
    response = HttpResponse(content_type=f'application/{tipo_doc}')  

    buff = BytesIO()  

    doc = SimpleDocTemplate(buff,  
        pagesize=tamaño_pagina,  
        rightMargin=40,  
        leftMargin=40,  
        topMargin=60,  
        bottomMargin=18,  
    ) 
    
    data = []  
    styles = getSampleStyleSheet()  
    styles = styles['Heading1']
    styles.alignment = TA_CENTER 

    header = Paragraph(f"{nombre}", styles)  
    
    data.append(header)  

    t = Table(lista)  

    t.setStyle(TableStyle(  
        [  
        ('GRID', (0, 0), (12, -1), 1, colors.dodgerblue),  
        ('LINEBELOW', (0, 0), (-1, 0), 3, colors.darkblue),  
        ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue)  
        ]  
    ))  
    
    data.append(t)

    if visitas:

        styles = getSampleStyleSheet()  
        styles.pageBreakBefore = 2
        styles = styles['Heading1']
        styles.alignment = TA_CENTER 
        header = Paragraph("Visitas", styles)  
        
        data.append(header)  

        t = Table(visitas)  

        t.setStyle(TableStyle(  
            [  
            ('GRID', (0, 0), (12, -1), 1, colors.dodgerblue),  
            ('LINEBELOW', (0, 0), (-1, 0), 3, colors.darkblue),  
            ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue)  
            ]  
        ))  
        
        data.append(t)

    doc.build(data)  
    response.write(buff.getvalue())   

    buff.seek(0)
    return FileResponse(buff, as_attachment=valor, filename=f'{nombre}.{extension}')

#*********************************************************************************

#********************************Pago fiados***************************************
@login_required(login_url="login")
def PagosFiadosListado(request):

    pagosF = PAGO_FIADO.objects.all()

    return render(request, 'pago_fiado/listar.html', {'pagosF':pagosF})

def DesactivarPagoFiado(request, id):
    pago_fiado = PAGO_FIADO.objects.get(id = id)
    pago_fiado.estado = 0
    pago_fiado.save()
    messages.warning(request, f'Pago fiado {pago_fiado.id} desactivado')
    return redirect('pagoFiadosListar')

def ActivarPagoFiado(request, id):
    pago_fiado = PAGO_FIADO.objects.get(id = id)
    pago_fiado.estado = 1
    pago_fiado.save()
    messages.warning(request, f'Pago fiado {pago_fiado.id} activado')
    return redirect('pagoFiadosListar')

@login_required(login_url="login")
def PagarFiado(request, id = None):
    
    mostrar = ''
    if id != None:
        mostrar = 'si'
    else:
        mostrar = 'no'

    print(mostrar)

    pagos_fiados = []
    pagos_fiados2 = []
    for pag in PAGO_FIADO.objects.all():
        pagos_fiados.append(pag.id)
        pagos_fiados.append(pag.cliente.nombre)
        pagos_fiados.append(pag.monto)
        pagos_fiados.append(pag.fecha_creacion)
        pagos_fiados2.append(pagos_fiados)
        pagos_fiados = []

    for i in DETALLE_FIADO.objects.all():
        for j in pagos_fiados2:
            if i.pago_fiado.id == j[0]:
                monto = j[2] - i.monto_abonado
                j[2] = monto

    pagos_fiados = []
    for obj in pagos_fiados2:
        if obj[0] == id:
            pagos_fiados.append(obj)

    detalleF = None
    if id == None:
        pagoF = pagos_fiados2

    else: 
        pagoF = pagos_fiados
        detalleF = DETALLE_FIADO.objects.filter(pago_fiado=id)

    if request.method == 'POST':

        pagoRealizado = request.POST.get('pago')

        pagoFiado = PAGO_FIADO.objects.get(id=id)

        pago_detalle = DETALLE_FIADO.objects.create(
            monto_abonado = pagoRealizado,
            pago_fiado = pagoFiado
        )
        pago_detalle.save()

        messages.warning(request, 'Pago realizado con exito')
        return redirect('PagarFiadoTodos')

    context = {
        'pagoF':pagoF,
        'detalleF':detalleF,
        'mostrar':mostrar
    }

    return render(request, 'pago_fiado/pagar_fiado.html', context)
